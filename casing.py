import json
import os
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                             QLabel, QLineEdit, QTextEdit, QFileDialog, QMessageBox,
                             QGroupBox, QStatusBar, QComboBox, QGridLayout, QTabWidget)
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import Qt
from docx import Document
import openpyxl
from HAD import HADCalculator

class DbCalculator(QWidget):
    def __init__(self):
        super().__init__()
        self.data_file = 'casing_data.json'
        self.had_data = {}
        self.had_calculator = HADCalculator()
        self.calculated_values = []
        self.first_at_head_value = None
        self.additional_info = []
        self.initUI()
        self.load_saved_data()

    def initUI(self):
        self.setStyleSheet(self.get_stylesheet())
        main_layout = QVBoxLayout()
        self.tab_widget = QTabWidget()
        input_tab = QWidget()
        results_tab = QWidget()
        self.tab_widget.addTab(input_tab, "Input")
        self.tab_widget.addTab(results_tab, "Results")
        
        input_layout = QVBoxLayout(input_tab)
        input_layout.addWidget(self.create_file_group())
        input_layout.addWidget(self.create_input_group())
        input_layout.addWidget(self.create_section_group())
        
        calculate_button = QPushButton("Calculate")
        calculate_button.setIcon(QIcon("icons/calculate.png"))
        calculate_button.clicked.connect(self.extract_and_display)
        input_layout.addWidget(calculate_button)
        
        results_layout = QHBoxLayout(results_tab)
        results_left_layout = QVBoxLayout()
        results_right_layout = QVBoxLayout()

        self.result_text = QTextEdit()
        self.result_text.setReadOnly(True)
        results_left_layout.addWidget(QLabel("Calculation Results:"))
        results_left_layout.addWidget(self.result_text)

        results_right_layout.addWidget(self.had_calculator)

        results_layout.addLayout(results_left_layout)
        results_layout.addLayout(results_right_layout)
        
        main_layout.addWidget(self.tab_widget)

        self.status_bar = QStatusBar()
        self.status_bar.showMessage("Ready")
        main_layout.addWidget(self.status_bar)

        save_button = QPushButton("Save Data")
        save_button.setIcon(QIcon("icons/save.png"))
        save_button.clicked.connect(self.save_data)
        main_layout.addWidget(save_button)

        self.setLayout(main_layout)

    def get_stylesheet(self):
        return """
            QWidget {
                background-color: #2b2b2b;
                color: #ffffff;
                font-family: 'Roboto', sans-serif;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 5px 15px;
                border-radius: 3px;
                margin: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QLineEdit, QTextEdit, QComboBox {
                background-color: #3b3b3b;
                border: 1px solid #555555;
                border-radius: 3px;
                padding: 3px;
                margin: 2px;
            }
            QGroupBox {
                border: 1px solid #555555;
                border-radius: 5px;
                margin-top: 10px;
                padding: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px 0 3px;
            }
            QLabel {
                margin: 2px;
            }
            QTabWidget::pane {
                border: 1px solid #555555;
                border-radius: 5px;
            }
            QTabBar::tab {
                background-color: #3b3b3b;
                border: 1px solid #555555;
                border-bottom-color: #555555;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                padding: 5px;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background-color: #4CAF50;
            }
        """

    def create_file_group(self):
        file_group = QGroupBox("File Selection")
        file_layout = QHBoxLayout()
        self.file_entry = QLineEdit()
        browse_button = QPushButton("Browse")
        browse_button.setIcon(QIcon("icons/folder.png"))
        browse_button.clicked.connect(self.select_file)
        file_layout.addWidget(self.file_entry)
        file_layout.addWidget(browse_button)
        file_group.setLayout(file_layout)
        return file_group

    def create_input_group(self):
        input_group = QGroupBox("Calculation Parameters")
        input_layout = QGridLayout()
        input_layout.addWidget(QLabel("Initial Dcsg amount:"), 0, 0)
        self.dcsg_entry = QLineEdit()
        input_layout.addWidget(self.dcsg_entry, 0, 1)
        input_layout.addWidget(QLabel("Number of iterations:"), 1, 0)
        self.iterations_entry = QLineEdit()
        input_layout.addWidget(self.iterations_entry, 1, 1)
        input_group.setLayout(input_layout)
        return input_group

    def create_section_group(self):
        section_group = QGroupBox("Section Parameters")
        section_layout = QGridLayout()
        self.section_inputs = []
        for i, section in enumerate(["Surface", "Intermediate", "Production"]):
            section_layout.addWidget(QLabel(f"{section} Section:"), i*3, 0, 1, 4)
            section_layout.addWidget(QLabel("Multiplier:"), i*3+1, 0)
            multiplier_entry = QLineEdit()
            section_layout.addWidget(multiplier_entry, i*3+1, 1)
            section_layout.addWidget(QLabel("Metal Type:"), i*3+1, 2)
            metal_type_combo = QComboBox()
            metal_type_combo.addItems(['K-55', 'L-80', 'N-80', 'P-110', 'Q-125', 'T-95', 'C-90'])
            section_layout.addWidget(metal_type_combo, i*3+1, 3)
            section_layout.addWidget(QLabel("Depth:"), i*3+2, 0)
            depth_entry = QLineEdit()
            section_layout.addWidget(depth_entry, i*3+2, 1)
            self.section_inputs.append((multiplier_entry, metal_type_combo, depth_entry))
        section_group.setLayout(section_layout)
        return section_group

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select File", "", "Word and Excel files (*.docx *.xlsx);;All files (*.*)"
        )
        if file_path:
            self.file_entry.setText(file_path)

    def find_at_body_value(self, file_path, at_head_value):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        at_head_col = None
        at_body_col = None
        
        for row in sheet.iter_rows(values_only=True):
            row_text = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            if "at head" in row_text:
                at_head_col = row_text.index("at head")
            if "at body" in row_text:
                at_body_col = row_text.index("at body")
            if at_head_col is not None and at_body_col is not None:
                break
                
        if at_head_col is None or at_body_col is None:
            return None
                
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                row_at_head = float(row[at_head_col])
                if abs(row_at_head - float(at_head_value)) < 0.01:
                    at_body_value = str(row[at_body_col])
                    if ' ' in at_body_value:
                        at_body_value = at_body_value.split()[-1]
                    return at_body_value
            except (ValueError, TypeError):
                continue
        return None

    def extract_values_from_docx(self, file_path, dcsg_amount):
        doc = Document(file_path)
        tables = doc.tables
        if not tables:
            QMessageBox.information(self, "Info", "No tables found in the document.")
            return None
        for table in tables:
            header_row = None
            at_head_index = None
            at_body_index = None
            for row in table.rows:
                row_text = [cell.text.strip() for cell in row.cells]
                if "At head" in row_text and "At body" in row_text:
                    header_row = row_text
                    at_head_index = header_row.index("At head")
                    at_body_index = header_row.index("At body")
                    break
            if header_row:
                for row in table.rows[1:]:
                    row_text = [cell.text.strip() for cell in row.cells]
                    if row_text[at_body_index] == dcsg_amount:
                        try:
                            return float(row_text[at_head_index])
                        except ValueError:
                            continue
        QMessageBox.information(self, "Info", f"No matching Dcsg amount ({dcsg_amount}) found in the document.")
        return None

    def extract_values_from_xlsx(self, file_path, dcsg_amount):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        at_head_col = None
        at_body_col = None
        for row in sheet.iter_rows(values_only=True):
            row_text = [str(cell).strip() if cell is not None else "" for cell in row]
            if at_head_col is None or at_body_col is None:
                if "At head" in row_text:
                    at_head_col = row_text.index("At head")
                if "At body" in row_text:
                    at_body_col = row_text.index("At body")
            elif at_head_col is not None and at_body_col is not None:
                if row_text[at_body_col] == dcsg_amount:
                    try:
                        return float(row_text[at_head_col])
                    except ValueError:
                        continue
        QMessageBox.information(self, "Info", f"No matching Dcsg amount ({dcsg_amount}) found in the document.")
        return None

    def find_nearest_bit_size_and_internal_diameter(self, file_path, db_value):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        bit_size_col = None
        internal_diameter_col = None
        bit_size_variations = ["bit size", "bitsize", "bit_size"]
        internal_diameter_variations = ["internal diameter"]
        for row in sheet.iter_rows(values_only=True):
            row_text = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            if bit_size_col is None:
                for i, cell in enumerate(row_text):
                    if any(variation in cell for variation in bit_size_variations):
                        bit_size_col = i
                    if any(variation in cell for variation in internal_diameter_variations):
                        internal_diameter_col = i
            if bit_size_col is not None and internal_diameter_col is not None:
                break
        if bit_size_col is None or internal_diameter_col is None:
            return None, None
        bit_sizes = []
        internal_diameters = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                bit_size = float(row[bit_size_col])
                internal_diameter = float(row[internal_diameter_col])
                bit_sizes.append(bit_size)
                internal_diameters.append(internal_diameter)
            except (ValueError, TypeError):
                continue
        if bit_sizes:
            nearest_bit_size = min(bit_sizes, key=lambda x: abs(x - db_value))
            nearest_index = bit_sizes.index(nearest_bit_size)
            return nearest_bit_size, internal_diameters[nearest_index]
        return None, None

    def find_reference_from_xlsx(self, file_path, internal_diameter_value):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        at_head_col = None
        internal_diameter_col = None
        for row in sheet.iter_rows(values_only=True):
            row_text = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            if "at head" in row_text:
                at_head_col = row_text.index("at head")
            if "internal diameter" in row_text:
                internal_diameter_col = row_text.index("internal diameter")
            if at_head_col is not None and internal_diameter_col is not None:
                break
        if at_head_col is None or internal_diameter_col is None:
            return f"Internal Diameter: {internal_diameter_value}, At head: Columns not found", None
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                row_internal_diameter = float(row[internal_diameter_col])
                if abs(row_internal_diameter - internal_diameter_value) < 0.01:
                    at_head_value = row[at_head_col]
                    return f"Internal Diameter: {row_internal_diameter}, At head (Dcsg): {at_head_value}", at_head_value
            except (ValueError, TypeError):
                continue
        return f"Internal Diameter: {internal_diameter_value}, At head: Not found", None

    def extract_additional_info(self, file_path, at_head_value, metal_type):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        at_head_col = None
        external_pressure_col = None
        metal_type_col = None
        tensile_strength_col = None
        unit_weight_col = None
        for row in sheet.iter_rows(values_only=True):
            row_text = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            if "at head" in row_text:
                at_head_col = row_text.index("at head")
            if "external pressure mpa" in row_text:
                external_pressure_col = row_text.index("external pressure mpa")
            if "metal type" in row_text:
                metal_type_col = row_text.index("metal type")
            if "tensile strength at body tonf" in row_text:
                tensile_strength_col = row_text.index("tensile strength at body tonf")
            if "unit weight length lbs/ft" in row_text:
                unit_weight_col = row_text.index("unit weight length lbs/ft")
            if all(col is not None for col in [at_head_col, external_pressure_col, metal_type_col, tensile_strength_col, unit_weight_col]):
                break
        if any(col is None for col in [at_head_col, external_pressure_col,metal_type_col, tensile_strength_col, unit_weight_col]):
            return []
        matching_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                row_at_head = float(row[at_head_col])
                row_metal_type = str(row[metal_type_col]).strip()
                if abs(row_at_head - float(at_head_value)) < 0.01 and row_metal_type == metal_type:
                    external_pressure = row[external_pressure_col]
                    tensile_strength = row[tensile_strength_col]
                    unit_weight = row[unit_weight_col]
                    matching_rows.append((row_at_head, external_pressure, row_metal_type, tensile_strength, unit_weight))
            except (ValueError, TypeError):
                continue
        return matching_rows

    def display_results(self, iteration, section, multiplier, metal_type, dcsg, db_value, nearest_bit_size, internal_diameter):
        at_body_value = self.find_at_body_value(self.file_entry.text(), dcsg)

        # Convert to inches (1 mm = 0.03937 inches)
        bit_size_inches = nearest_bit_size * 0.03937 if nearest_bit_size else None
        dcsg_inches = float(dcsg) * 0.03937 if dcsg else None
        at_body_inches = float(at_body_value) * 0.03937 if at_body_value else None

        # Format the values consistently
        formatted_bit_size = f"{nearest_bit_size:.2f} mm ({bit_size_inches:.2f}\")" if nearest_bit_size else "-"
        formatted_dcsg = f"{dcsg} mm ({dcsg_inches:.2f}\")" if dcsg else "-"
        formatted_at_body = f"{at_body_value} mm ({at_body_inches:.2f}\")" if at_body_value else "-"
        formatted_internal_diameter = f"{internal_diameter:.2f}" if internal_diameter else "-"

        if iteration == 1:
            self.result_text.clear()
            header = f"""
            <div style='margin-top: 20px; margin-bottom: 20px;'>
                <h3 style='color: #4CAF50; padding: 10px;'>Calculation Results</h3>
                <table style='border-collapse: collapse; width: 100%; background-color: #2b2b2b;'>
                    <tr style='background-color: #4CAF50; color: white;'>
                        <th style='padding: 8px; text-align: center; border: 1px solid #555; width: 20%;'>Section</th>
                        <th style='padding: 8px; text-align: center; border: 1px solid #555; width: 20%;'>Nearest Bit Size</th>
                        <th style='padding: 8px; text-align: center; border: 1px solid #555; width: 20%;'>DCSG</th>
                        <th style='padding: 8px; text-align: center; border: 1px solid #555; width: 20%;'>DCSG'</th>
                        <th style='padding: 8px; text-align: center; border: 1px solid #555; width: 20%;'>Internal Diameter</th>
                    </tr>
            """
            self.result_text.setHtml(header)

        bg_color = "#3b3b3b" if iteration % 2 == 0 else "#2b2b2b"
        row_html = f"""
            <tr style='background-color: {bg_color};'>
                <td style='padding: 8px; text-align: center; color: white; border: 1px solid #555; width: 20%;'>{section}</td>
                <td style='padding: 8px; text-align: center; color: white; border: 1px solid #555; width: 20%;'>{formatted_bit_size}</td>
                <td style='padding: 8px; text-align: center; color: white; border: 1px solid #555; width: 20%;'>{formatted_dcsg}</td>
                <td style='padding: 8px; text-align: center; color: white; border: 1px solid #555; width: 20%;'>{formatted_at_body}</td>
                <td style='padding: 8px; text-align: center; color: white; border: 1px solid #555; width: 20%;'>{formatted_internal_diameter}</td>
            </tr>
        """

        current_html = self.result_text.toHtml()
        if "</table>" in current_html:
            current_html = current_html.replace("</table>", row_html + "</table>")
        else:
            current_html += row_html + "</table></div>"
        
        self.result_text.setHtml(current_html)

    def calculate_had(self, depth, matching_rows, section_name):
        if section_name != "Production Section":
            return False
        
        s_values = {
            'K-55': 1.05,
            'L-80': 1.08,
            'N-80': 1.08,
            'P-110': 1.125,
            'Q-125': 1.125,
            'T-95': 1.125,
            'C-90': 1.125
        }
        for row in matching_rows:
            at_head, external_pressure, metal_type, tensile_strength, unit_weight = row
            s = s_values.get(metal_type, 1.08)
            ep = float(external_pressure)
            had = (100 * ep) / (s * 1.08)
            self.display_had_results(at_head, external_pressure, metal_type, had, depth, tensile_strength, unit_weight, section_name)
            if had >= depth:
                return True
        return False

    def display_had_results(self, at_head, external_pressure, metal_type, had, depth, tensile_strength, unit_weight, section_name):
        at_head_key = round(at_head, 2)
        if at_head_key not in self.had_data:
            self.had_data[at_head_key] = []
        self.had_data[at_head_key].append({
            'had': had,
            'external_pressure': external_pressure,
            'metal_type': metal_type,
            'tensile_strength': tensile_strength,
            'unit_weight': unit_weight
        })
        if had >= depth:
            self.had_calculator.update_had_results(self.had_data, depth, section_name)

    def extract_and_display(self):
        file_path = self.file_entry.text()
        initial_dcsg_amount = self.dcsg_entry.text()
        try:
            iterations = int(self.iterations_entry.text())
        except ValueError:
            QMessageBox.critical(self, "Error", "Please enter a valid number of iterations.")
            return
        if not file_path:
            QMessageBox.critical(self, "Error", "Please select a file first.")
            return
        if not initial_dcsg_amount:
            QMessageBox.critical(self, "Error", "Please enter an initial Dcsg amount.")
            return

        self.result_text.clear()
        self.had_calculator.had_text.clear()
        self.had_data.clear()

        try:
            dcsg_amount = initial_dcsg_amount
            at_head_value = None
            self.calculated_values = []
            self.first_at_head_value = None
            next_dcsg = None
            self.additional_info = []

            for i in range(min(iterations, 3)):
                try:
                    multiplier = float(self.section_inputs[i][0].text())
                except ValueError:
                    QMessageBox.critical(self, "Error", f"Please enter a valid multiplier for {['Production', 'Intermediate', 'Surface'][i]} section.")
                    return

                metal_type = self.section_inputs[i][1].currentText()
                try:
                    depth = float(self.section_inputs[i][2].text())
                except ValueError:
                    QMessageBox.critical(self, "Error", f"Please enter a valid depth for {['Production', 'Intermediate', 'Surface'][i]} section.")
                    return

                if i == 0:
                    if file_path.lower().endswith('.docx'):
                        at_head_value = self.extract_values_from_docx(file_path, dcsg_amount)
                    elif file_path.lower().endswith('.xlsx'):
                        at_head_value = self.extract_values_from_xlsx(file_path, dcsg_amount)
                    else:
                        QMessageBox.critical(self, "Error", "Unsupported file format.")
                        return

                    if at_head_value is not None:
                        self.first_at_head_value = at_head_value
                        dcsg_amount = str(at_head_value)
                    else:
                        self.result_text.append("First iteration - At head value not found")
                        return

                if at_head_value is not None:
                    db_value = float(at_head_value) * multiplier
                    nearest_bit_size, internal_diameter = self.find_nearest_bit_size_and_internal_diameter(file_path, db_value)

                    if nearest_bit_size is not None and internal_diameter is not None:
                        self.display_results(
                            i+1,
                            ['Production', 'Intermediate', 'Surface'][i],
                            multiplier,
                            metal_type,
                            dcsg_amount,
                            db_value,
                            nearest_bit_size,
                            internal_diameter
                        )

                        reference_result, new_at_head_value = self.find_reference_from_xlsx(file_path, internal_diameter)
                        next_dcsg = new_at_head_value
                        self.calculated_values.append((at_head_value, db_value, nearest_bit_size))
                        matching_rows = self.extract_additional_info(file_path, at_head_value, metal_type)
                        self.additional_info.extend(matching_rows)
                        
                        section_name = ['Production', 'Intermediate', 'Surface'][i] + " Section"
                        if section_name == "Production Section":
                            if not self.calculate_had(depth, matching_rows, section_name):
                                self.result_text.append(f"Could not find a suitable HAD value for the given depth in {section_name}.")
                                break

                        if i < min(iterations, 3) - 1:
                            if new_at_head_value is not None:
                                at_head_value = new_at_head_value
                                dcsg_amount = str(new_at_head_value)
                            else:
                                self.result_text.append("Could not find a new 'At head' value. Stopping iterations.")
                                break
                        else:
                            self.calculated_values.append((new_at_head_value, None, None))
                            if new_at_head_value is not None:
                                self.result_text.append("")
                    else:
                        self.result_text.append("Bit Size and Internal Diameter columns not found or empty.")
                        break
                else:
                    self.result_text.append("No matching Dcsg amount found.")
                    break

        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")

        self.status_bar.showMessage("Calculation completed")
        self.tab_widget.setCurrentIndex(1)

    def save_data(self):
        data = {
            'file_path': self.file_entry.text(),
            'initial_dcsg': self.dcsg_entry.text(),
            'iterations': self.iterations_entry.text(),
            'section_inputs': [
                {
                    'multiplier': input[0].text(),
                    'metal_type': input[1].currentText(),
                    'depth': input[2].text()
                }
                for input in self.section_inputs
            ]
        }
        with open(self.data_file, 'w') as f:
            json.dump(data, f)
        self.status_bar.showMessage("Data saved successfully", 3000)

    def load_saved_data(self):
        if os.path.exists(self.data_file):
            with open(self.data_file, 'r') as f:
                data = json.load(f)
            self.file_entry.setText(data.get('file_path', ''))
            self.dcsg_entry.setText(data.get('initial_dcsg', ''))
            self.iterations_entry.setText(data.get('iterations', ''))
            
            section_inputs = data.get('section_inputs', [])
            for i, input_data in enumerate(section_inputs):
                if i < len(self.section_inputs):
                    self.section_inputs[i][0].setText(input_data.get('multiplier', ''))
                    self.section_inputs[i][1].setCurrentText(input_data.get('metal_type', 'K-55'))
                    self.section_inputs[i][2].setText(input_data.get('depth', ''))

    def get_all_dcsg_values(self):
        initial_dcsg = self.dcsg_entry.text()
        at_head_values = [value[0] for value in self.calculated_values if value[0] is not None]
        nearest_bit_sizes = [value[2] for value in self.calculated_values if value[2] is not None]
        return initial_dcsg, at_head_values, nearest_bit_sizes, None
